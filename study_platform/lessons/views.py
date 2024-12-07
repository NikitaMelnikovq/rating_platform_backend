import io
import base64
from datetime import timedelta

from django.urls import reverse
from django.shortcuts import get_object_or_404
from django.http import HttpResponse
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.permissions import IsAuthenticated
from rest_framework.views import APIView
from rest_framework import generics, filters, status
from rest_framework.response import Response
import qrcode 
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from django.utils.timezone import localtime


from .serializers import LessonSerializer, StudentFeedbackSerializer
from .models import Lesson, StudentFeedback
from accounts.models import User
from .tasks import process_feedback
from accounts.pagination import Pagination

class LessonCreateView(generics.CreateAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = LessonSerializer

    def perform_create(self, serializer):
        lesson = serializer.save(teacher=self.request.user)
        feedback_url = self.request.build_absolute_uri(
            reverse('lessons:student-feedback', kwargs={'unique_code': lesson.unique_code})
        )

        img = qrcode.make(feedback_url)
        buffer = io.BytesIO()
        img.save(buffer, format='PNG')
        qr_code_base64 = base64.b64encode(buffer.getvalue()).decode('utf-8')
        lesson.qr_code_base64 = qr_code_base64
        lesson.save()
        self.qr_code_base64 = qr_code_base64

    def create(self, request, *args, **kwargs):
        response = super().create(request, *args, **kwargs)
        response.data['qr_code'] = self.qr_code_base64
        return response
    
    def get_serializer_context(self):
        context = super().get_serializer_context()
        context.update({"request": self.request})
        return context

class LessonDeleteView(APIView):
    permission_classes = [IsAuthenticated]

    def delete(self, request, unique_code):
        try:
            lesson = Lesson.objects.get(unique_code=unique_code, teacher=request.user)
            lesson.delete()
            return Response({"message": "Lesson deleted successfully."}, status=status.HTTP_204_NO_CONTENT)
        except Lesson.DoesNotExist:
            return Response({"error": "Lesson not found."}, status=status.HTTP_404_NOT_FOUND)


class LessonDetailView(generics.RetrieveUpdateAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = LessonSerializer
    queryset = Lesson.objects.all()
    lookup_field = 'unique_code'

    def get_serializer_context(self):
        context = super().get_serializer_context()
        context.update({"request": self.request})
        return context

    def get_queryset(self):
        return Lesson.objects.filter(teacher=self.request.user)

    def patch(self, request, *args, **kwargs):
        return self.partial_update(request, *args, **kwargs)

class LessonStatView(generics.RetrieveAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = LessonSerializer

    def get_serializer_context(self):
        context = super().get_serializer_context()
        context.update({"request": self.request})
        return context

    def get_object(self):
        unique_code = self.kwargs.get("unique_code")

        return get_object_or_404(Lesson, unique_code=unique_code)


class LessonByCodeView(generics.RetrieveAPIView):
    serializer_class = LessonSerializer
    lookup_field = 'unique_code'
    queryset = Lesson.objects.all()

    def get_serializer_context(self):
        context = super().get_serializer_context()
        context.update({"request": self.request})
        return context

    def get(self, request, *args, **kwargs):
        lesson = self.get_object()
        if not lesson.is_link_active():
            return Response({'error': 'Link is not active'}, status=status.HTTP_400_BAD_REQUEST)
        serializer = self.get_serializer(lesson)
        return Response(serializer.data)


class TeacherLessonListView(generics.ListAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = LessonSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['start_time', 'end_time']
    search_fields = ['topic']

    def get_serializer_context(self):
        context = super().get_serializer_context()
        context.update({"request": self.request})
        return context

    def get_queryset(self):
        return Lesson.objects.filter(teacher=self.request.user)
    


class TeacherLessonListByIdView(generics.ListAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = LessonSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    pagination_class = Pagination
    filterset_fields = ['start_time', 'end_time']
    search_fields = ['topic']
    ordering_fields = ['average_rating', 'start_time']

    def get_queryset(self):
        return Lesson.objects.filter(teacher_id=self.kwargs['id'])

    def list(self, request, *args, **kwargs):
        response = super().list(request, *args, **kwargs)
        teacher_id = self.kwargs['id']
        teacher = User.objects.get(id=teacher_id, role='teacher')
        feedbacks = StudentFeedback.objects.filter(lesson__teacher=teacher)
        good_counts = {}
        bad_counts = {}

        for fb in feedbacks:
            if not fb.praises:
                continue
            if fb.rating >= 3:
                for p in fb.praises:
                    good_counts[p] = good_counts.get(p, 0) + 1
            else:
                for p in fb.praises:
                    bad_counts[p] = bad_counts.get(p, 0) + 1

        most_frequent_praise = max(good_counts, key=good_counts.get) if good_counts else None
        most_frequent_criticism = max(bad_counts, key=bad_counts.get) if bad_counts else None

        response.data['teacher'] = {
            'id': teacher.id,
            'role': teacher.role,
            'first_name': teacher.first_name,
            'surname': teacher.surname,
            'last_name': teacher.last_name or '',
            'rating': teacher.rating,
            'feedback_count': teacher.feedback_count,
            'most_frequent_praise': most_frequent_praise,
            'most_frequent_criticism': most_frequent_criticism
        }

        return response

class StudentFeedbackCreateView(generics.CreateAPIView):
    serializer_class = StudentFeedbackSerializer

    def post(self, request, *args, **kwargs):
        lesson_code = kwargs.get('unique_code')
        try:
            lesson = Lesson.objects.get(unique_code=lesson_code)
            if not lesson.is_link_active():
                return Response({'error': 'Link is not active'}, status=status.HTTP_400_BAD_REQUEST)
            
            feedback_data = request.data.copy()
            feedback_data['lesson'] = lesson.id

            process_feedback.delay(feedback_data)

            return Response({'message': 'Feedback received and will be processed shortly.'}, status=status.HTTP_202_ACCEPTED)

        except Lesson.DoesNotExist:
            return Response({'error': 'Invalid lesson code'}, status=status.HTTP_404_NOT_FOUND)

class IncreaseTimeView(APIView):
    permission_classes = [IsAuthenticated]

    def patch(self, request, unique_code):
        try:
            lesson = Lesson.objects.get(unique_code=unique_code, teacher=request.user)

            lesson.end_time += timedelta(minutes=10)
            lesson.save()

            return Response(
                {"message": "Activation duration increased by 10 minutes."},
                status=status.HTTP_200_OK,
            )
        except Lesson.DoesNotExist:
            return Response(
                {"error": "Lesson not found or you don't have permission."},
                status=status.HTTP_404_NOT_FOUND,
            )
        

def generate_qr_code(request, unique_code):
    lesson = get_object_or_404(Lesson, unique_code=unique_code)
    frontend_base_url = "http://localhost:5173"  
    url = f"{frontend_base_url}/form/{lesson.unique_code}/"
    img = qrcode.make(url)
    response = HttpResponse(content_type="image/png")
    img.save(response, "PNG")
    return response



class TeacherExcelReportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, id):
        teacher = User.objects.get(id=id, role='teacher')
        feedbacks = StudentFeedback.objects.filter(lesson__teacher=teacher)

        good_counts = {}
        bad_counts = {}

        for fb in feedbacks:
            if not fb.praises:
                continue
            if fb.rating >= 4:
                for p in fb.praises:
                    good_counts[p] = good_counts.get(p, 0) + 1
            else:
                for p in fb.praises:
                    bad_counts[p] = bad_counts.get(p, 0) + 1

        most_frequent_praise = max(good_counts, key=good_counts.get) if good_counts else None
        most_frequent_criticism = max(bad_counts, key=bad_counts.get) if bad_counts else None

        institute_name = teacher.institute.name if teacher.institute else ''
        total_reviews = teacher.feedback_count
        rating = teacher.rating or 0.0

        wb = Workbook()

        ws1 = wb.active
        ws1.title = "Основная информация"

        ws1.append(["ФИО преподавателя", "Рейтинг", "Институт", "Топ 2 мини-отзыва", "Количество оценок"])

        fio = f"{teacher.first_name or ''} {teacher.surname or ''} {teacher.last_name or ''}".strip()

        top_reviews = []
        if most_frequent_praise:
            top_reviews.append(most_frequent_praise)
        if most_frequent_criticism:
            top_reviews.append(most_frequent_criticism)

        row_values = [fio, rating, institute_name, ", ".join(top_reviews), total_reviews]
        ws1.append(row_values)

        top_reviews_cell = ws1.cell(row=2, column=4)
        parts = top_reviews_cell.value.split(", ") if top_reviews_cell.value else []

        ws1.delete_cols(4)
        ws1.cell(row=1, column=4, value="Похвала")
        ws1.cell(row=1, column=5, value="Критика")
        ws1.cell(row=1, column=6, value="Количество оценок")

        ws1.cell(row=2, column=1, value=fio)
        ws1.cell(row=2, column=2, value=rating)
        ws1.cell(row=2, column=3, value=institute_name)
        ws1.cell(row=2, column=6, value=total_reviews)

        praise_cell = ws1.cell(row=2, column=4, value=most_frequent_praise if most_frequent_praise else '')
        criticism_cell = ws1.cell(row=2, column=5, value=most_frequent_criticism if most_frequent_criticism else '')

        green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
        red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

        if most_frequent_praise:
            praise_cell.fill = green_fill
        if most_frequent_criticism:
            criticism_cell.fill = red_fill

        ws2 = wb.create_sheet("Отзывы")

        # Добавляем ещё одну колонку после "Результат" для praises
        ws2.append(["Студент", "Предмет", "Пара (Тема)", "Дата и время", "Комментарий", "Оценка", "Результат", "Praises"])

        teacher_lessons = Lesson.objects.filter(teacher=teacher)
        feedbacks = StudentFeedback.objects.filter(lesson__in=teacher_lessons).select_related('lesson', 'lesson__subject')

        for fb in feedbacks:
            dt = localtime(fb.created_at).strftime('%Y-%m-%d %H:%M:%S')
            result_text = "Понравилось" if fb.rating >= 4 else "Не понравилось"
            praises_str = ", ".join(fb.praises) if fb.praises else ''
            ws2.append([
                fb.student_name,
                fb.lesson.subject.name if fb.lesson.subject else '',
                fb.lesson.topic,
                dt,
                fb.comment,
                fb.rating,
                result_text,
                praises_str
            ])

        from io import BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"report_teacher_{teacher.id}.xlsx"

        response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

